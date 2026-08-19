from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO

from flask import redirect, flash, request, url_for
from flask_login import current_user
from .email import send_email

from collections import defaultdict
from lxml import etree

from sqlalchemy import and_, or_
from sqlalchemy.sql import func, or_
import math

from common_models.src import current_utc_time
from .models import Organization, Report, Sections, Version_report
from . import db

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ZERO_DECIMAL = Decimal('0.00')

def parse_int(value):
    try:
        return int(value)
    except (ValueError, TypeError):
        return None

def to_decimal(value):
    if not value and value != 0:
        return Decimal('0.00')
    try:
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        cleaned = str(value).replace(',', '.').strip()
        return Decimal(cleaned)
    except (ValueError, InvalidOperation, TypeError, AttributeError):
        return Decimal('0.00')

def check_version_editable(version):
    if not version:
        flash('Версия не найдена.', 'error')
        return False
    if version.status in ('Отправлен', 'Одобрен'):
        flash('Редактирование отправленного/одобренного отчета недоступно.', 'error')
        return False
    return True

def update_version_status(version):
    if version:
        version.change_time = current_utc_time()
        version.status = "Заполнение"
        db.session.commit()

def create_section(data, product_id, code_product):
    return Sections(
        id_version=data['current_version_id'],
        id_product=product_id,
        code_product=code_product,
        section_number=data['section_number'],
        Oked=data['oked'],
        produced=data['produced'],
        Consumed_Quota=data['Consumed_Quota'],
        Consumed_Fact=data['Consumed_Fact'],
        Consumed_Total_Quota=data['Consumed_Total_Quota'],
        Consumed_Total_Fact=data['Consumed_Total_Fact'],
        total_differents=None,
        note=data['note']
    )

def calculate_consumed_fact(section, product_unit):
    if section.produced == 0:
        return Decimal('0.00')
    if product_unit and product_unit.NameUnit in ('%', '% (включая покупную)'):
        value = (section.Consumed_Total_Fact / section.produced) * 100
    else:
        value = (section.Consumed_Total_Fact / section.produced) * 1000
    
    return Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

def calculate_total_quota(section, product_unit):
    if section.Consumed_Quota == 0:
        return Decimal('0.00')
    if product_unit and product_unit.NameUnit in ('%', '% (включая покупную)'):
        value = (section.produced * section.Consumed_Quota) / 100
    else:
        value = (section.produced * section.Consumed_Quota) / 1000
    
    return Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

def process_section_calculations(section, product_unit):
    try:
        if section.produced != 0:
            section.Consumed_Fact = calculate_consumed_fact(section, product_unit)
        else:
            section.Consumed_Fact = Decimal('0.00')
        
        section.Consumed_Total_Quota = calculate_total_quota(section, product_unit)
        section.total_differents = section.Consumed_Total_Fact - section.Consumed_Total_Quota
        db.session.commit()
    except InvalidOperation as e:
        flash(f"Ошибка при вычислениях: {e}")

def update_aggregated_sections(version_id, section_number):
    specific_codes = ['9001', '9010', '9100']
    
    section9001 = Sections.query.filter_by(
        id_version=version_id, section_number=section_number, code_product='9001'
    ).first()
    
    aggregated = db.session.query(
        func.sum(Sections.Consumed_Total_Quota),
        func.sum(Sections.Consumed_Total_Fact),
        func.sum(Sections.total_differents)
    ).filter(
        Sections.id_version == version_id,
        Sections.section_number == section_number,
        ~Sections.code_product.in_(specific_codes)
    ).first()
    
    if section9001 and aggregated:
        section9001.Consumed_Total_Quota = aggregated[0] or 0
        section9001.Consumed_Total_Fact = aggregated[1] or 0
        section9001.total_differents = aggregated[2] or 0
    
    section9010 = Sections.query.filter_by(
        id_version=version_id, section_number=section_number, code_product='9010'
    ).first()
    section9100 = Sections.query.filter_by(
        id_version=version_id, section_number=section_number, code_product='9100'
    ).first()
    
    if section9100 and section9001 and section9010:
        section9100.Consumed_Total_Quota = section9001.Consumed_Total_Quota + section9010.Consumed_Total_Quota
        section9100.Consumed_Total_Fact = section9001.Consumed_Total_Fact + section9010.Consumed_Total_Fact
        section9100.total_differents = section9001.total_differents + section9010.total_differents
    
    db.session.commit()

def update_section_fields(section, form, product_unit):
    if section.product.CodeProduct == "7000":
        section.Consumed_Total_Quota = to_decimal(form.get('Consumed_Total_Quota_change'))
        section.Consumed_Total_Fact = to_decimal(form.get('Consumed_Total_Fact_change'))
        section.note = form.get('note_change')
        db.session.commit()
        section.total_differents = section.Consumed_Total_Fact - section.Consumed_Total_Quota
    else:
        section.produced = to_decimal(form.get('produced_change'))
        section.Consumed_Quota = to_decimal(form.get('Consumed_Quota_change'))
        section.Consumed_Total_Fact = to_decimal(form.get('Consumed_Total_Fact_change'))
        section.note = form.get('note_change')
        db.session.commit()
        process_section_calculations(section, product_unit)
    db.session.commit()

def subtract_from_aggregated_sections(section):
    for code in ['9001', '9100']:
        agg_section = Sections.query.filter_by(
            id_version=section.id_version,
            section_number=section.section_number,
            code_product=code
        ).first()
        
        if agg_section:
            agg_section.Consumed_Total_Quota -= section.Consumed_Total_Quota
            agg_section.Consumed_Total_Fact -= section.Consumed_Total_Fact
            agg_section.total_differents -= section.total_differents
    
    db.session.commit()

def redirect_back(version, section_number=None):
    if section_number:
        report_type_map = {1: 'fuel', 2: 'heat', 3: 'electro'}
        report_type = report_type_map.get(section_number)
        if report_type:
            return redirect(url_for('views.report_section', report_type=report_type, id=version.id))
    return redirect(request.referrer)

def cancel_sending(id):
    current_version = Version_report.query.filter_by(id=id).first()
    if current_version.status == 'Отправлен':
        current_version.status = 'Заполнение'
        db.session.commit()
        flash('Отправка отчета была отменена! Новый статус отчета - "Заполнение".', 'succes')
    else:
        flash('Отменить можно только непросмотренный отчет!', 'error')
    return redirect(request.referrer)

def get_organizations_with_reports_excel_xlsx(year: int, quarter: int, statuses: list) -> bytes:
    status_filter = 'all_reports' if not statuses else statuses[0] if len(statuses) == 1 else 'all_reports'
    
    reports = get_reports_by_status(status_filter, year, quarter)
    
    if not reports:
        return None

    organizations_data = set()
    for report in reports:
        valid_versions = [v for v in report.versions if not statuses or v.status in statuses]
        if valid_versions:
            latest_version = max(valid_versions, key=lambda x: x.sent_time or datetime.min)
            organizations_data.add((
                report.organization.okpo,
                report.organization.full_name,
                latest_version.sent_time
            ))

    if not organizations_data:
        return None

    records = sorted(organizations_data, key=lambda x: x[0] or "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Организации"

    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill("solid", fgColor="C6EFCE")
    title_font = Font(bold=True, size=12)
    title_fill = PatternFill("solid", fgColor="D9E1F2")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    start_row = 3
    start_col = 2

    title_text = f"Список предприятий, представивших отчеты по форме Сведения о нормах в электронном виде на {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
    ws.merge_cells(start_row=start_row - 1, start_column=start_col, end_row=start_row - 1, end_column=start_col + 3)
    title_cell = ws.cell(row=start_row - 1, column=start_col, value=title_text)
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[start_row - 1].height = 45

    for col in range(start_col, start_col + 4):
        cell = ws.cell(row=start_row - 1, column=col)
        cell.border = thin_border

    headers = ['Код предприятия (ОКПО)', 'Наименование предприятия', 'Дата поступления', 'Примечание']
    for i, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    for row_offset, (okpo, full_name, sent_time) in enumerate(records, start=1):
        row_data = [okpo or '', full_name or '', sent_time or '', '']
        for col_offset, value in enumerate(row_data):
            cell = ws.cell(
                row=start_row + row_offset,
                column=start_col + col_offset,
                value=value
            )
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
            if col_offset == 2 and sent_time:
                cell.number_format = 'YYYY-MM-DD'

    for i in range(len(headers)):
        col_letter = get_column_letter(start_col + i)
        max_length = len(headers[i])
        for j in range(1, len(records) + 1):
            val = ws.cell(row=start_row + j, column=start_col + i).value
            if val:
                max_length = max(max_length, len(str(val)))
        ws.column_dimensions[col_letter].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def control_func(id):
    current_version = Version_report.query.filter_by(id=id).first()
    if not current_version:
        flash('Версия отчета не найдена.', 'error')
        return redirect(url_for('views.report_area'))

    id_version = current_version.id
    sections = {
        'fuel': Sections.query.filter_by(id_version=id, section_number=1, code_product='9010').first(),
        'heat': Sections.query.filter_by(id_version=id, section_number=2, code_product='9010').first(),
        'electro': Sections.query.filter_by(id_version=id, section_number=3, code_product='9010').first(),
    }

    if current_version.status == 'Заполнение':
        for key, section in sections.items():
            if section is None or not section.note:
                flash('«Примечание» с кодом строки 9010 обязательно для заполнения.', 'error')
                return redirect(url_for('views.report_section', report_type=key, id=id_version))

        current_version.status = 'Контроль пройден'
        db.session.commit()
        flash('Контроль пройден.', 'successful')
    else:
        flash('Контроль уже был пройден.', 'error')

    return redirect(request.referrer) 

def get_reports_by_status(status, year=None, quarter=None, region=None):
    def translate_status(status):
        status_map = {
            'not_viewed': 'Отправлен',
            'remarks': 'Есть замечания',
            'to_download': 'Одобрен',
            'to_delete': 'Готов к удалению'
        }
        return status_map.get(status)

    filters = []
    statuses = [
        'Отправлен',
        'Есть замечания',
        'Одобрен',
        'Готов к удалению'
    ]
    
    if year:
        filters.append(Report.year == year)
    if quarter:
        filters.append(Report.quarter == quarter)
    
    user_type = current_user.type
    
    region_filter = None
    if region and region.isdigit():
        region_filter = int(region)

    if user_type == "Администратор" or user_type == "Смотрящий":
        base_query = Report.query.join(Version_report).join(Organization)
        
        if region_filter:
            base_query = base_query.filter(Organization.region.has(number=region_filter))
        
        if status == 'all_reports':
            query = base_query.filter(
                or_(*[Version_report.status == s for s in statuses]),
                *filters
            )
        else:
            trans_status = translate_status(status)
            if trans_status:
                query = base_query.filter(
                    Version_report.status == trans_status,
                    *filters
                )
            else:
                return []
        return query.order_by(Version_report.sent_time.desc()).all()
    
    else:
        if current_user.organization and current_user.organization.region:
            user_region_number = current_user.organization.region.number
        else:
            user_region_number = None
        
        if status == 'all_reports':
            query = Report.query.join(Version_report).join(Organization)
            
            if user_region_number:
                query = query.filter(Organization.region.has(number=user_region_number))
            
            return query.filter(
                or_(*[Version_report.status == s for s in statuses]),
                *filters
            ).order_by(Version_report.sent_time.desc()).all()
            
        else:
            trans_status = translate_status(status)
            if trans_status:
                query = Report.query.join(Version_report).join(Organization)
                
                if user_region_number:
                    query = query.filter(Organization.region.has(number=user_region_number))
                
                return query.filter(
                    Version_report.status == trans_status,
                    *filters
                ).order_by(Version_report.sent_time.desc()).all()
            else:
                return []